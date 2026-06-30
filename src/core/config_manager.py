"""
RailPway_Technical

Module:
    config_manager.py

Purpose:
    Load and manage application configuration files.

Author:
    Sirshendu Karmakar

Version:
    0.1.0
"""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.logger import logger


class ConfigManager:
    """
    Loads and provides access to handbook configuration.
    """

    def __init__(self, project_root: Path):

        self.project_root = Path(project_root)

        self.handbook_config = (
            self.project_root
            / "handbook"
            / "config"
            / "handbook_config.xlsx"
        )

        self.chapter_index = (
            self.project_root
            / "handbook"
            / "config"
            / "chapter_index.xlsx"
        )

        self.config: dict[str, Any] = {}

    # ---------------------------------------------------------

    def load_handbook_config(self) -> None:
        """
        Load handbook_config.xlsx
        """

        logger.info("Loading handbook configuration...")

        if not self.handbook_config.exists():
            raise FileNotFoundError(self.handbook_config)

        wb = load_workbook(self.handbook_config, data_only=True)

        ws = wb["Configuration"]

        for row in ws.iter_rows(min_row=2, values_only=True):

            key = row[0]
            value = row[1]

            self.config[key] = value

        logger.info("Configuration loaded successfully.")

    # ---------------------------------------------------------

    def get(self, key: str, default=None):

        return self.config.get(key, default)

    # ---------------------------------------------------------

    def show(self):

        logger.info("Current Configuration")

        for k, v in self.config.items():

            logger.info("%s : %s", k, v)