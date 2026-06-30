"""
RailPway_Technical

Module:
    excel_manager.py

Purpose:
    Read and write Excel workbooks for the handbook.

Author:
    Sirshendu Karmakar

Version:
    0.1.0
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from core.logger import logger


class ExcelManager:
    """
    Handles Excel workbook operations.
    """

    def __init__(self):

        self.workbook = None
        self.worksheet = None

    # ---------------------------------------------------------

    def open(self, file_path: Path, sheet_name: str | None = None):

        logger.info("Opening workbook : %s", file_path)

        if not file_path.exists():
            raise FileNotFoundError(file_path)

        self.workbook = load_workbook(file_path)

        if sheet_name:

            self.worksheet = self.workbook[sheet_name]

        else:

            self.worksheet = self.workbook.active

        logger.info("Workbook opened successfully.")

    # ---------------------------------------------------------

    def create(self):

        logger.info("Creating new workbook")

        self.workbook = Workbook()

        self.worksheet = self.workbook.active

    # ---------------------------------------------------------

    def save(self, file_path: Path):

        logger.info("Saving workbook : %s", file_path)

        self.workbook.save(file_path)

    # ---------------------------------------------------------

    def read_cell(self, row: int, column: int):

        return self.worksheet.cell(row=row, column=column).value

    # ---------------------------------------------------------

    def write_cell(self, row: int, column: int, value: Any):

        self.worksheet.cell(row=row, column=column).value = value

    # ---------------------------------------------------------

    def get_all_rows(self):

        for row in self.worksheet.iter_rows(values_only=True):
            yield row

    # ---------------------------------------------------------

    def append(self, values: list):

        self.worksheet.append(values)

    # ---------------------------------------------------------

    def max_row(self):

        return self.worksheet.max_row

    # ---------------------------------------------------------

    def max_column(self):

        return self.worksheet.max_column